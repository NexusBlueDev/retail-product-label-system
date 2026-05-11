"""
ls_13mwz_bigfamily_upc_fix.py
================================
Fixes UPC codes and adds CUSTOM SKU codes to the 108 matched variants
in the big 700-variant 13MWZ Navy family (c3c968b4).

Source of truth: docs/Wrangler 13MWZ.xlsx (Corrinne's corrected import file)
Target: product-export (11).xlsx (199 currently-visible LS variants)
Match key: Color + Size + Length (variant option values)

Action per matched variant:
  PUT v2.1 /products/{id}
  {"details": {"product_codes": [
      {"type": "CUSTOM", "code": <Corrinne's NEW SKU>},
      {"type": "UPC",    "code": <corrected 12-digit UPC>}
  ]}}

The 91 unmatched RIGID variants are written to: docs/ls_13mwz_unmatched_rigid.csv
for Corrinne's review.

Usage:
  python3 docs/ls_13mwz_bigfamily_upc_fix.py [--dry-run]
"""

import sys
import csv
import json
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

DOCS = Path("docs")
DRY_RUN = "--dry-run" in sys.argv
LS_BASE_V20 = "https://therodeoshop.retail.lightspeed.app/api/2.0"
LS_BASE_V21 = "https://therodeoshop.retail.lightspeed.app/api/2.1"
LOG_FILE = DOCS / "ls_13mwz_bigfamily_fix.log"
UNMATCHED_CSV = DOCS / "ls_13mwz_unmatched_rigid.csv"

# ── credentials ────────────────────────────────────────────────────────────────

def get_ls_token() -> str:
    env = Path(".env.local").read_text()
    for line in env.splitlines():
        if line.startswith("LIGHTSPEED_TOKEN="):
            return line.split("=", 1)[1].strip()
    raise RuntimeError("LIGHTSPEED_TOKEN not found in .env.local")

# ── LS API ─────────────────────────────────────────────────────────────────────

def ls_put(token: str, variant_id: str, payload: dict) -> tuple[dict, int]:
    url = f"{LS_BASE_V21}/products/{variant_id}"
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, method="PUT", headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "curl/7.81.0",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read()), resp.status
    except urllib.error.HTTPError as e:
        body = {}
        try:
            body = json.loads(e.read())
        except Exception:
            pass
        return body, e.code

def ls_get(token: str, variant_id: str) -> tuple[dict, int]:
    url = f"{LS_BASE_V20}/products/{variant_id}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": "curl/7.81.0",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read()), resp.status
    except urllib.error.HTTPError as e:
        return {}, e.code

# ── data loading ───────────────────────────────────────────────────────────────

def load_corrinne_lookup() -> dict:
    """Build lookup: (color, size, length) -> {upc, new_sku}"""
    wb = openpyxl.load_workbook(DOCS / "Wrangler 13MWZ.xlsx")
    ws = wb.active
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        color   = str(row[15]).strip() if row[15] else ""
        size    = str(row[18]).strip() if row[18] else ""
        length  = str(row[20]).strip() if row[20] else ""
        upc     = str(row[12]).strip() if row[12] else ""
        new_sku = str(row[10]).strip() if row[10] else ""
        if len(upc) == 11:
            upc = "0" + upc
        if color and size and length:
            lookup[(color, size, length)] = {"upc": upc, "new_sku": new_sku}
    return lookup

def load_export_variants() -> list[dict]:
    """Load 199 LS-visible variants from the product export."""
    wb = openpyxl.load_workbook(DOCS / "product-export (11).xlsx")
    ws = wb["Product Export"]
    headers = [str(c.value) for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

# ── merge logic ────────────────────────────────────────────────────────────────

def build_update_list(export_variants: list[dict], corrinne: dict) -> tuple[list, list]:
    """Returns (matched, unmatched) variant lists."""
    matched = []
    unmatched = []
    for v in export_variants:
        color  = str(v.get("variant_option_one_value")   or "").strip()
        size   = str(v.get("variant_option_two_value")   or "").strip()
        length = str(v.get("variant_option_three_value") or "").strip()
        key = (color, size, length)
        if key in corrinne:
            matched.append({
                "id":      v["id"],
                "sku":     v["sku"],
                "color":   color,
                "size":    size,
                "length":  length,
                "new_sku": corrinne[key]["new_sku"],
                "upc":     corrinne[key]["upc"],
            })
        else:
            unmatched.append({
                "id":    v["id"],
                "sku":   v["sku"],
                "color": color,
                "size":  size,
                "length": length,
            })
    return matched, unmatched

# ── main ───────────────────────────────────────────────────────────────────────

def main():
    token = get_ls_token()
    corrinne = load_corrinne_lookup()
    export   = load_export_variants()
    matched, unmatched = build_update_list(export, corrinne)

    print(f"13MWZ Big-Family UPC Fix — {'DRY RUN' if DRY_RUN else 'LIVE'}")
    print(f"Started: {datetime.now(timezone.utc).isoformat()}")
    print(f"Matched:   {len(matched)} variants to update")
    print(f"Unmatched: {len(unmatched)} variants (no Corrinne file entry)\n")

    # Write unmatched report
    with open(UNMATCHED_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "sku", "color", "size", "length"])
        w.writeheader()
        w.writerows(unmatched)
    print(f"Unmatched variants written to: {UNMATCHED_CSV}\n")

    if DRY_RUN:
        print("DRY RUN — first 5 planned updates:")
        for v in matched[:5]:
            print(f"  {v['id'][:8]} {v['sku']:20s} {v['color']:20s} {v['size']}x{v['length']} -> UPC={v['upc']} CUSTOM={v['new_sku']}")
        print("  ...")
        print("\nRun without --dry-run to apply.")
        return

    # Execute updates
    stats = {"ok": 0, "skipped": 0, "error": 0}
    log_lines = []

    for i, v in enumerate(matched, 1):
        vid = v["id"]

        # GET current product_codes to avoid wiping unknown existing codes
        current, get_status = ls_get(token, vid)
        if get_status != 200:
            msg = f"GET failed ({get_status}) for {vid}"
            print(f"  [{i}/{len(matched)}] ERROR: {msg}")
            log_lines.append(f"ERROR\t{vid}\t{v['sku']}\t{msg}")
            stats["error"] += 1
            time.sleep(0.5)
            continue

        # Extract existing non-UPC codes to preserve them
        existing_codes = []
        for c in (current.get("data", {}) or {}).get("product_codes", []):
            code_type = c.get("type", "")
            if code_type not in ("UPC",):
                existing_codes.append({"type": code_type, "code": c.get("code", "")})

        # Build new product_codes: existing non-UPC + new CUSTOM (replaces old) + new UPC
        new_codes = [
            {"type": "CUSTOM", "code": v["new_sku"]},
            {"type": "UPC",    "code": v["upc"]},
        ]
        # Keep any non-CUSTOM, non-UPC codes that might exist
        for c in existing_codes:
            if c["type"] not in ("CUSTOM", "UPC"):
                new_codes.append(c)

        payload = {"details": {"product_codes": new_codes}}
        result, put_status = ls_put(token, vid, payload)

        if put_status in (200, 201):
            stats["ok"] += 1
            log_lines.append(f"OK\t{vid}\t{v['sku']}\t{v['color']}\t{v['size']}x{v['length']}\tUPC={v['upc']}\tCUSTOM={v['new_sku']}")
            if i % 10 == 0:
                print(f"  [{i}/{len(matched)}] {stats['ok']} ok, {stats['error']} errors")
        else:
            err_msg = str(result)[:200]
            stats["error"] += 1
            log_lines.append(f"ERROR\t{vid}\t{v['sku']}\t{err_msg}")
            print(f"  [{i}/{len(matched)}] ERROR: {v['sku']} → {put_status}: {err_msg}")

        time.sleep(0.3)  # ~3 req/sec

    # Write log
    with open(LOG_FILE, "w") as f:
        f.write("\n".join(log_lines))

    print(f"\nDone: {stats['ok']} updated, {stats['error']} errors")
    print(f"Log: {LOG_FILE}")
    print(f"Unmatched (91 RIGID): {UNMATCHED_CSV}")

if __name__ == "__main__":
    main()
